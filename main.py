import os
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes
from openai import OpenAI
from openpyxl import load_workbook
from datetime import datetime

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

FILE_NAME = "finance.xlsx"

def init_excel():
    try:
        load_workbook(FILE_NAME)
    except:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Type", "Amount", "Category", "Note"])
        wb.save(FILE_NAME)

init_excel()

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    response = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {"role": "system", "content": "Extract finance data: type (Income/Expense), amount (number), category, note. Respond in format: type,amount,category,note"},
            {"role": "user", "content": text}
        ]
    )

    result = response.choices[0].message.content

    try:
        t, amount, category, note = result.split(",")

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        ws.append([
            datetime.now().strftime("%Y-%m-%d"),
            t,
            int(amount),
            category,
            note
        ])

        wb.save(FILE_NAME)

        await update.message.reply_text(f"Saved ✅ {t} {amount} {category}")

    except:
        await update.message.reply_text("Error 😅 Try: Spent 50k on food")

app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

app.run_polling()
